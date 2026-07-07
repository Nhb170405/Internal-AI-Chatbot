from openpyxl import load_workbook

from app.ingestion.parsed_document import ParsedDocument
from app.ingestion.text_cleanup import cleanup_extracted_text


def parse_xlsx(file_path: str, file_name: str) -> ParsedDocument:
    # Muc tieu:
    # 1. Doc workbook xlsx.
    # 2. Lap qua tung sheet.
    # 3. Dung Table-Like Simple Parser: Sheet -> Row -> Column.
    # 4. Khong doan header phuc tap, khong tach nhieu bang trong sheet.
    #
    # Chua can xu ly style, color, merged cell trong Milestone 5.
    workbook = load_workbook(file_path, data_only=True, read_only=True)

    lines = [
        f"Document: {file_name}",
        "File type: XLSX",
        "Strategy: raw sheet/row/cell matrix",
        "",
    ]

    sheet_count = 0
    sheet_names: list[str] = []
    total_rows = 0
    max_column_count = 0

    for sheet in workbook.worksheets:
        sheet_has_data = False

        for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            cleaned_cells = normalize_row_values(row)

            if not any(cleaned_cells):
                continue

            if not sheet_has_data:
                lines.append(f"Sheet: {sheet.title}")
                lines.append("")
                sheet_has_data = True
                sheet_count += 1
                sheet_names.append(sheet.title)

            lines.append(f"Row {row_index}:")

            for column_index, cell in enumerate(cleaned_cells, start=1):
                if cell:
                    lines.append(f"Column {column_index}: {cell}")

            lines.append("")
            total_rows += 1
            max_column_count = max(max_column_count, len(cleaned_cells))

    if total_rows == 0:
        raise ValueError("XLSX file is empty.")

    extracted_text = "\n".join(lines)
    cleaned_text = cleanup_extracted_text(extracted_text)

    if not cleaned_text:
        raise ValueError("XLSX file is empty.")

    return ParsedDocument(
        text=cleaned_text,
        parser_name="xlsx_parser",
        character_count=len(cleaned_text),
        page_count=None,
        metadata={
            "strategy": "raw_sheet_row_cell_matrix",
            "sheetCount": sheet_count,
            "sheets": sheet_names,
            "totalRows": total_rows,
            "maxColumnCount": max_column_count,
        },
        warnings=[],
    )


def normalize_row_values(row: tuple[object, ...]) -> list[str]:
    # Muc tieu:
    # 1. Chuyen moi cell value ve string de dua vao extracted text.
    # 2. Cell None thanh chuoi rong.
    # 3. Strip khoang trang dau/cuoi.
    cleaned_cells: list[str] = []

    for value in row:
        if value is None:
            cleaned_cells.append("")
        else:
            cleaned_cells.append(str(value).strip())

    return cleaned_cells
